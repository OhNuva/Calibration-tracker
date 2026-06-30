# calibration tracker - anuva nath
# started this after ortech bc we were tracking everything on paper and it was annoying
# someone else had to retype all of it into excel manually which seems insane
# this just does it automatically and tells you whats overdue

import csv
import os
from datetime import date, datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# TODO: maybe add email alerts later? would be useful for the lab manager


def load_equipment(filepath):
    equipment = []
    with open(filepath, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # need to convert the date string to an actual date object or the math doesnt work
            row["last_calibrated"] = datetime.strptime(row["last_calibrated"], "%Y-%m-%d").date()
            row["calibration_interval_days"] = int(row["calibration_interval_days"])
            equipment.append(row)
    return equipment


def get_status(equipment, warn_days=14):
    today = date.today()
    results = []

    for item in equipment:
        next_due = item["last_calibrated"] + timedelta(days=item["calibration_interval_days"])
        days_left = (next_due - today).days

        # negative = already passed the due date
        if days_left < 0:
            status = "OVERDUE"
        elif days_left <= warn_days:
            status = "DUE SOON"
        else:
            status = "OK"

        results.append({
            "id":              item["equipment_id"],
            "name":            item["name"],
            "location":        item["location"],
            "last_cal":        item["last_calibrated"],
            "interval":        item["calibration_interval_days"],
            "next_due":        next_due,
            "days_left":       days_left,
            "status":          status,
            "responsible":     item["responsible"],
            "notes":           item.get("notes", ""),
        })

    return results


def print_report(results):
    overdue  = [r for r in results if r["status"] == "OVERDUE"]
    soon     = [r for r in results if r["status"] == "DUE SOON"]
    fine     = [r for r in results if r["status"] == "OK"]

    print("\n-------------------------------------------")
    print("  calibration status - " + str(date.today()))
    print("-------------------------------------------")

    print(f"\nOVERDUE ({len(overdue)})")
    if len(overdue) == 0:
        print("  none!")
    for r in overdue:
        print(f"  {r['id']} | {r['name']} | {abs(r['days_left'])} days overdue | {r['responsible']}")
        if r["notes"]:
            print(f"       note: {r['notes']}")

    print(f"\nDUE SOON within 14 days ({len(soon)})")
    if len(soon) == 0:
        print("  none!")
    for r in soon:
        print(f"  {r['id']} | {r['name']} | due in {r['days_left']} days ({r['next_due']}) | {r['responsible']}")

    print(f"\nOK ({len(fine)})")
    for r in fine:
        print(f"  {r['id']} | {r['name']} | due in {r['days_left']} days")

    print(f"\n-------------------------------------------")
    print(f"total: {len(results)} | overdue: {len(overdue)} | due soon: {len(soon)} | ok: {len(fine)}")
    print("-------------------------------------------\n")


def make_excel(results, output="calibration_report.xlsx"):

    # colour coding - red overdue, yellow coming up, green fine
    red    = PatternFill(start_color="FFFF6B6B", end_color="FFFF6B6B", fill_type="solid")
    yellow = PatternFill(start_color="FFFFD966", end_color="FFFFD966", fill_type="solid")
    green  = PatternFill(start_color="FFB6D7A8", end_color="FFB6D7A8", fill_type="solid")
    grey   = PatternFill(start_color="FF434343", end_color="FF434343", fill_type="solid")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "calibration status"

    headers = ["id", "equipment name", "location", "last calibrated", "interval (days)", "next due", "days left", "status", "responsible", "notes"]

    # header row
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = grey
        c.font = Font(bold=True, color="FFFFFFFF")
        c.alignment = Alignment(horizontal="center")

    # sort so overdue shows up first
    sorted_r = sorted(results, key=lambda x: x["days_left"])

    for row_i, r in enumerate(sorted_r, 2):
        data = [
            r["id"],
            r["name"],
            r["location"],
            r["last_cal"].strftime("%Y-%m-%d"),
            r["interval"],
            r["next_due"].strftime("%Y-%m-%d"),
            r["days_left"],
            r["status"],
            r["responsible"],
            r["notes"],
        ]

        if r["status"] == "OVERDUE":
            row_fill = red
        elif r["status"] == "DUE SOON":
            row_fill = yellow
        else:
            row_fill = green

        for col_i, val in enumerate(data, 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.fill = row_fill
            c.alignment = Alignment(horizontal="center")

    # make columns not tiny
    for col in ws.columns:
        max_w = max((len(str(cell.value)) for cell in col if cell.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_w + 3

    ws.freeze_panes = "A2"

    wb.save(output)
    print(f"saved to {output}")


# runs everything
equipment = load_equipment("equipment.csv")
results = get_status(equipment)
print_report(results)
make_excel(results)

